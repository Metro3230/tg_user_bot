from decouple import config
import os
from datetime import datetime
from pyrogram import Client, filters, enums
from openpyxl import load_workbook, Workbook
import schedule
import time
from pathlib import Path
import asyncio
from pyrogram.types import Message
from pyrogram.raw.functions.messages import GetDialogFilters
from pyrogram.raw.types import DialogFilter, DialogFilterDefault
import logging


app = Client(name=config('LOGIN'),
             api_id=config('API_ID'),
             api_hash=config('API_HASH'),
             phone_number=config('PHONE'))


script_dir = Path(__file__).parent  # Определяем путь к текущему скрипту
excel_path = script_dir / "chats.xlsx"
log_file = script_dir / 'log.log'


#  логгер для моего скрипта
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)  #  уровень логов для моего скрипта

#  обработчик для записи в файл
file_handler = logging.FileHandler(log_file)
file_handler.setLevel(logging.INFO)  # Уровень для файлового обработчика
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

#  логгер для сторонних библиотек
logging.getLogger().setLevel(logging.WARNING)





def update_chat_list():
    if os.path.exists(excel_path):    # Открываем или создаем Excel-файл
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID чата", "Название", "Дата добавления"])  # Заголовки

# ------коньструкция выхода из чатов по времени, кроме исключений--------------
    try:
        current_date = datetime.now()        # Устанавливаем текущую дату для сравнения

        chat_ids_to_leave = []        # массив на удаление
        chat_ids_to_exceptions = []    # массив с исключениями
        filtered_chats = []   # списки на удаленние - списки в "Исключения"
        delay_days = int(ws['D1'].value)  #кол-во дней задержки перед выходом

        dialog_filters = app.invoke(GetDialogFilters()) # наполнение массива исключений из папки "Исключения"
        for filter in dialog_filters:
            if isinstance(filter, DialogFilter):
                if filter.title == 'Исключения' or filter.title == 'исключения':
                    for pinned_peers in filter.pinned_peers:
                        if hasattr(pinned_peers, 'channel_id'):
                            chat_ids_to_exceptions.append(pinned_peers.channel_id)
                        if hasattr(pinned_peers, 'user_id'):
                            chat_ids_to_exceptions.append(pinned_peers.user_id)
                    for include_peers in filter.include_peers:
                        if hasattr(include_peers, 'channel_id'):
                            chat_ids_to_exceptions.append(include_peers.channel_id)
                        if hasattr(include_peers, 'user_id'):
                            chat_ids_to_exceptions.append(include_peers.user_id)

        # Проверяем таблицу на условия выхода из чатов
        for row in ws.iter_rows(min_row=2, max_col=4):  # min_row=2, чтобы пропустить заголовки
            chat_id, chat_title, added_date = row[0].value, row[1].value, row[2].value

            # Проверка на необходимость выхода по дате
            if added_date:
                added_date_dt = datetime.strptime(added_date, "%Y-%m-%d %H:%M:%S")
                if (current_date - added_date_dt).days > delay_days and check_substring_in_array(chat_id, chat_ids_to_exceptions) == False:   # проверка на время и исключение
                    chat_ids_to_leave.append((chat_id, chat_title))

        # Выход из чатов 
    except Exception as e:
        logger.error(f'ошибка выхода из чатов по времени: {e}')

    if len(chat_ids_to_leave) > 0:
        app.send_message("me", f"Скрипт выходит из следующих чатов, так как в них уже более {delay_days} дней:")
        for chat_id, chat_title in chat_ids_to_leave:   #  выйти из чата и вывести его название
            try:
                app.leave_chat(chat_id)
                app.send_message("me", f"{chat_title}")
            except Exception as e:
                app.send_message("me", f"Ошибка при выходе из {chat_title} : {e}")
# ------------------------------------------------------------------------------

# ------коньструкция удаления из таблички чатов в которых аккаунта больше нет---
    try:
        rows_to_delete = []

        # Получаем список текущих диалогов и сразу преобразуем его в список
        dialogs = list(app.get_dialogs())
        current_chat_ids = {dialog.chat.id for dialog in dialogs}

        # Удаляем строки с чатами, которые больше не существуют в аккаунте
        for row in ws.iter_rows(min_row=2, max_col=1):  # min_row=2, чтобы пропустить заголовки
            chat_id = row[0].value
            if chat_id not in current_chat_ids:
                rows_to_delete.append(row[0].row)

        # Удаление строк для чатов, из которых вышли, или которых больше нет в аккаунте
        for row in reversed(rows_to_delete):  # Удаляем строки с конца, чтобы избежать сдвига индексов
            ws.delete_rows(row)
    except Exception as e:
        logger.error(f'ошибка удаления из таблички чатов в которых аккаунта больше нет: {e}')
# -------------------------------------------------------------------------------

# ------коньструкция добавления в таболичку новых аккаунтов----------------------
    try:
        # Получаем список всех существующих Chat IDs из обновленной таблицы
        existing_chat_ids = {row[0].value for row in ws.iter_rows(min_row=2, max_col=1)}

        # Добавляем только новые чаты
        new_entries = []
        for dialog in dialogs:
            chat = dialog.chat
            if chat.id not in existing_chat_ids:
                new_entry = [
                    chat.id,
                    chat.title or chat.first_name or "Bot",
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ""  # Пустое значение в столбце "Action" по умолчанию
                ]
                new_entries.append(new_entry)
                ws.append(new_entry)  # Добавляем новую строку в таблицу
# -------------------------------------------------------------------------------

# -------------охраняем файл, если были добавлены или удалены строки-------------
        if new_entries or rows_to_delete:
            wb.save(excel_path)
            logger.info(f"{len(new_entries)} новых чатов добавлено, {len(rows_to_delete)} чатов удалено из таблицы.")
        else:
            logger.info('Изменений в списке чатов не найдено.')
    except Exception as e:
        logger.error(f'ошибка добавления в таболичку новых аккаунтов: {e}')
# ---------------------------------------------------------------------------------



# ------фугкция проверки входжлдения строки в массив строк----------------------
def check_substring_in_array(target_string, array):
    target_string = str(target_string)
    # Проверяем, содержит ли целевая строка хотя бы одну из строк массива
    for item in array:
        if str(item) in target_string:
            return True  # Совпадение найдено, возвращаем True
    return False  # Если совпадений не было найдено
# ---------------------------------------------------------------------------------

# ------фугкция обновыления времени задержки до выхода----------------------
def write_delay(msg):

    try:
        command_parts = msg.split(maxsplit=2)         # Разделяем текст команды на части

        if len(command_parts) < 2:         # Проверяем, что есть и пароль
            app.send_message("me", "Ошибка: формат команды /write_delаy <дней>")
            return
        
        number = command_parts[1]

        # Открываем или создаем Excel-файл
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["ID чата", "Название", "Дата добавления"])  # Заголовки

        ws['D1'] = number    # Записываем число в ячейку D1

        wb.save(excel_path)    # Сохраняем изменения
        app.send_message("me", f"Чаты (группы, паблики и т.д.) будут удаляться спустя {number} дней. (всё в папке \"Исключения\" останенся на месте.)")
    except Exception as e:
        app.send_message("me", f"Произошла ошибка: {e}")
# ---------------------------------------------------------------------------------




####

####

####


def check_for_command():
    messages = app.get_chat_history("me", limit=1)    # Получаем последние сообщения в Избранном
    for message in messages:

        if message.text == "/help":
            wb = load_workbook(excel_path)
            ws = wb.active
            delay_days = ws['D1'].value  #кол-во дней задержки перед выходом
            app.send_message("me",  f"Скрипт работает. Выходит из групп с датами вступления старше {delay_days} дней.\n\n"
                                    "/write_delаy <дней> - установить количество дней, после истечении которых - вы выйдете из пабликов и чатов (за исключением папки \"Исключения\").\n")

        elif message.text == "/service":
            app.send_message("me",  "/write_delаy <дней> - установить количество дней, после истечении которых - паблики и чаты (за исключением папки \"Исключения\") удаляться.\n"
                                    "/dw - скачать актуальную табличку\n"
                                    "/update - обновить принудительно\n"
                                    "/log - скачать логи\n"
                                    "можно еще скинуть табличку с именем chats.xlsx в чятик и она заменить актуальную там, но хз зачем это, лучше не баловаться")

        elif "/write_delay" in message.text:
            write_delay(message.text)

        elif message.text == "/log":
            if os.path.exists(log_file):            # Отправляем файл в Избранное
                app.send_document("me", log_file)
                app.send_message("me", "/\ логи.")
            else:
                app.send_message("me", "Файл не найден.")

        elif message.text == "/dw":
            if os.path.exists(excel_path):            # Отправляем файл в Избранное
                app.send_document("me", excel_path)
                app.send_message("me", "/\ табличка чатов.")
            else:
                app.send_message("me", "Файл не найден.")

        elif message.text == "/update":
            app.send_message("me", "Обновляю..")
            update_chat_list()

        elif message.document and message.document.file_name == "chats.xlsx":            # Проверяем, является ли сообщение документом с именем "chats.xlsx"
            app.download_media(message.document, file_name=excel_path)
            app.send_message("me", f"Файл '{message.document.file_name}' успешно сохранен на сервере.")
            break  # Завершаем проверку после сохранения файла




# Планирование обновления таблицы раз в сутки в час ночи
schedule.every().day.at("01:00").do(update_chat_list)

logger.info(f"скрипт запущен")

# Запуск приложения и планировщика
with app:
    update_chat_list()
    while True:
        schedule.run_pending()        
        check_for_command()        # Проверяем наличие сообщений в избранном
        time.sleep(10)  # Проверка раз 10 сек

